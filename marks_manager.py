from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


NAME_COL = 1
DOB_COL = 2
BLOODGROUP_COL = 3
AGE_COL = 4

FIRST_SUBJECT_COL = 5
LAST_SUBJECT_COL = 10

TOTAL_SCORE_COL = 11
AVERAGE_COL = 12
RANK_COL = 13
GRADE_COL = 14

REMARKS_COL = 15

EXPECTED_HEADERS = [
    "Name", "D.O.B", "Blood group", "Age",
    "Tamil", "English", "Maths", "Physics", "Chemistry", "Computer science",
    "Total", "Average", "Rank", "Grade", "Remarks"
]


def load_workbook_and_sheet(path, sheet_name):
    """Load an Excel workbook and return the requested worksheet.

    Args:
        path (str): Path to the Excel file.
        sheet_name (str): Name of the worksheet to load.

    Returns:
        tuple: (workbook, worksheet) if successful, (None, None) if file not found or sheet doesn't exist.
    """
    try:
        wb = load_workbook(path)
    except FileNotFoundError:
        print(f"The file {path} is not found")
        return None, None

    if sheet_name not in wb.sheetnames:
        print(f"The given sheetname {sheet_name} is not in {path}")
        print(f"The available sheets are {wb.sheetnames}")
        return None, None

    ws = wb[sheet_name]
    return wb, ws


def validate_headers(ws):
    """Validate that worksheet headers match expected column names.

    Args:
        ws: The worksheet object to validate.

    Returns:
        bool: True if all headers match, False otherwise.
    """
    for col in range(1, len(EXPECTED_HEADERS) + 1):
        expected = EXPECTED_HEADERS[col - 1]
        actual = ws[get_column_letter(col) + '1'].value

        if actual != expected:
            print(f"There is a mismatch in the headers in column no: {col},")
            print(f"The expected header name is {expected}")
            print(f"The actual header is {actual}")
            return False
    return True


def recalculate_all(ws):
    """Recalculate total marks, average, grades and remarks for all students.

    Args:
        ws: The worksheet object containing student data.

    Returns:
        bool: True if calculation completed successfully.
    """
    for row in range(2, ws.max_row + 1):
        current_name = ws[get_column_letter(NAME_COL) + str(row)].value
        if current_name is None:
            continue

        total = 0
        for col in range(FIRST_SUBJECT_COL, LAST_SUBJECT_COL + 1):
            total += ws[get_column_letter(col) + str(row)].value

        ws[get_column_letter(TOTAL_SCORE_COL) + str(row)] = total
        avg_cal = total / (LAST_SUBJECT_COL - FIRST_SUBJECT_COL + 1)
        ws[get_column_letter(AVERAGE_COL) + str(row)] = avg_cal

        if avg_cal >= 90:
            ws[get_column_letter(GRADE_COL) + str(row)] = "A1"
        elif avg_cal >= 80:
            ws[get_column_letter(GRADE_COL) + str(row)] = "A2"
        elif avg_cal >= 70:
            ws[get_column_letter(GRADE_COL) + str(row)] = "B1"
        elif avg_cal >= 60:
            ws[get_column_letter(GRADE_COL) + str(row)] = "B2"
        elif avg_cal >= 50:
            ws[get_column_letter(GRADE_COL) + str(row)] = "C1"
        elif avg_cal >= 40:
            ws[get_column_letter(GRADE_COL) + str(row)] = "C2"
        else:
            ws[get_column_letter(GRADE_COL) + str(row)] = "F"

        low_subjects = []
        for col in range(FIRST_SUBJECT_COL, LAST_SUBJECT_COL + 1):
            subject_name = ws[get_column_letter(col) + '1'].value
            if ws[get_column_letter(col) + str(row)].value < 40:
                low_subjects.append(subject_name)

        if low_subjects:
            ws[get_column_letter(REMARKS_COL) + str(row)] = \
                "Needs improvement in " + ", ".join(low_subjects)
        elif avg_cal >= 75:
            ws[get_column_letter(REMARKS_COL) + str(row)] = "EXCELLENT PERFORMANCE"
        elif avg_cal >= 60:
            ws[get_column_letter(REMARKS_COL) + str(row)] = "Good, can improve further"
        else:
            ws[get_column_letter(REMARKS_COL) + str(row)] = "Needs more practice"

    return True


def calculate_ranks(ws):
    """Calculate ranks for all students based on total marks using Excel RANK.EQ function.

    Failed students (Grade = F) get 'FAIL' in rank column.
    Passing students get an Excel RANK.EQ formula that updates dynamically.

    Args:
        ws: The worksheet object containing student data.
    """
    total_col_letter = get_column_letter(TOTAL_SCORE_COL)
    total_range = f"${total_col_letter}$2:${total_col_letter}${ws.max_row}"

    for row in range(2, ws.max_row + 1):
        if ws[get_column_letter(NAME_COL) + str(row)].value is None:
            continue

        grade = ws[get_column_letter(GRADE_COL) + str(row)].value
        if grade == "F":
            ws[get_column_letter(RANK_COL) + str(row)].value = "FAIL"
        else:
            ws[get_column_letter(RANK_COL) + str(row)].value = \
                f"=RANK.EQ({total_col_letter}{row},{total_range},0)"


def entry_of_marks():
    """Get marks for one student from user input.

    Returns:
        list: List of 6 subject marks, or None if any mark is invalid (not 0-100).
    """
    student_marks = []
    print("Enter the student's marks in the order of Tamil, English, Maths, Physics, Chemistry, Computer science:")
    for i in range(6):
        val = int(input())
        if val < 0 or val > 100:
            print("The mark is invalid (marks should be 0 <= mark <= 100)")
            return None
        student_marks.append(val)

    return student_marks


def add_student_interactive(ws):
    """Add a new student record to the worksheet interactively.

    Prompts user for student name, DOB, blood group, age, and marks in 6 subjects.
    Automatically recalculates totals, grades, and ranks after adding.

    Args:
        ws: The worksheet object to add the student to.
    """
    next_row = ws.max_row + 1

    student_info = []
    student_info.append(input("Enter the name of the student : "))
    student_info.append(input("Enter the DOB in the format of dd-mm-yy : "))
    student_info.append(input("Enter the Blood Group : "))
    student_info.append(input("Enter the Age : "))

    marks = entry_of_marks()
    if marks is None or len(marks) != 6:
        print("The mark entry has failed, student is not added!")
        return

    # write basic info
    for col in range(1, AGE_COL + 1):
        ws[get_column_letter(col) + str(next_row)] = student_info[col - 1]

    # write marks
    i = 0
    for col in range(FIRST_SUBJECT_COL, LAST_SUBJECT_COL + 1):
        ws[get_column_letter(col) + str(next_row)] = marks[i]
        i += 1

    recalculate_all(ws)
    calculate_ranks(ws)
    ws.parent.save("students_template.xlsx")
    print(f"Student '{student_info[0]}' added and file saved successfully!")


def export_fail_students(ws, output_path):
    """Export all students with grade F to a new Excel file."""
    wb_new = Workbook()
    ws_new = wb_new.active

    ws_new.title = "Failed Students List"

    # copy headers
    for col in range(1, ws.max_column + 1):
        ws_new[get_column_letter(col) + '1'] = ws[get_column_letter(col) + '1'].value

    new_row = 2
    failed_count = 0

    for row in range(2, ws.max_row + 1):
        name = ws[get_column_letter(NAME_COL) + str(row)].value
        if name is None:
            continue

        grade = ws[get_column_letter(GRADE_COL) + str(row)].value
        if grade == "F":
            for col in range(1, ws.max_column + 1):
                ws_new[get_column_letter(col) + str(new_row)] = \
                    ws[get_column_letter(col) + str(row)].value
            new_row += 1
            failed_count += 1

    wb_new.save(output_path)
    print(f"Export completed: {failed_count} failed students exported to {output_path}")
    if failed_count == 0:
        print("Note: No failed students found. Make sure you ran 'Recalculate marks, grades, ranks' first.")


def main():
    wb, ws = load_workbook_and_sheet("students_template.xlsx", "Students Details")
    if ws is None:
        exit()

    if not validate_headers(ws):
        print("The headers are mismatched")
        exit()
    else:
        print("The headers are matched")

    while True:
        print("\n--- Student Marks Menu ---")
        print("1. Recalculate marks, grades, ranks")
        print("2. Add a new student")
        print("3. Export failed students")
        print("4. Save and exit")

        choice = input("Enter your choice (1-4): ").strip()

        if choice == "1":
            recalculate_all(ws)
            calculate_ranks(ws)
            print("Recalculated marks, grades and ranks.")

        elif choice == "2":
            add_student_interactive(ws)
            print("Finished add_student_interactive.")

        elif choice == "3":
            output_path = input("Enter output file name for failed students (e.g. failed_students.xlsx): ").strip()
            if output_path == "":
                output_path = "failed_students.xlsx"
            export_fail_students(ws, output_path)

        elif choice == "4":
            wb.save("students_template.xlsx")
            print("Workbook saved as students_template.xlsx. Exiting.")
            break

        else:
            print("Invalid choice. Please enter 1, 2, 3, or 4.")


if __name__ == "__main__":
    main()